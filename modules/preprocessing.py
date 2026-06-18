import re
import csv
import os
import json

try:
    from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
    from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
except ImportError:
    StemmerFactory = None
    StopWordRemoverFactory = None

_stemmer = None
_stopword_remover = None

def get_stemmer():
    global _stemmer
    if _stemmer is None and StemmerFactory is not None:
        factory = StemmerFactory()
        _stemmer = factory.create_stemmer()
    return _stemmer

def get_stopword_remover():
    global _stopword_remover
    if _stopword_remover is None and StopWordRemoverFactory is not None:
        factory = StopWordRemoverFactory()
        _stopword_remover = factory.create_stop_word_remover()
    return _stopword_remover

# BAGIAN 1: PEMBACAAN FILE (CSV & EXCEL)
def read_csv_file(
    filepath: str,
    delimiter: str = ','
) -> list[dict]:
    
    result = []

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File tidak ditemukan: {filepath}")

    with open(filepath, mode='r', encoding='utf-8-sig', newline='') as f:
        # csv.DictReader otomatis menjadikan baris pertama sebagai header (key dict)
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            # Konversi OrderedDict → dict biasa, strip whitespace per nilai
            result.append({k: v.strip() if isinstance(v, str) else v
                           for k, v in dict(row).items()})

    return result


def read_excel_file(
    filepath: str,
    sheet_name: str = None
) -> list[dict]:

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl diperlukan. Install: pip install openpyxl")

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File tidak ditemukan: {filepath}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Pilih worksheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    result = []
    headers = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Baris pertama (i=0) → jadikan header
        if i == 0:
            # Normalisasi header: lowercase & strip spasi
            headers = [str(cell).strip().lower() if cell is not None else f'col_{j}'
                       for j, cell in enumerate(row)]
            continue

        # Baris data: zip header dengan nilai sel
        # zip() → pasangkan setiap header[j] dengan row[j]
        row_values = [str(cell).strip() if cell is not None else ''
                      for cell in row]

        # Lewati baris yang seluruhnya kosong
        if all(v == '' for v in row_values):
            continue

        row_dict = dict(zip(headers, row_values))
        result.append(row_dict)

    wb.close()
    return result


def read_dataset(
    filepath: str,
    text_col: str = 'text',
    label_cols: list[str] = None
) -> list[dict]:

    if label_cols is None:
        label_cols = ['label']
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        raw = read_csv_file(filepath)
    elif ext in ('.xlsx', '.xls'):
        raw = read_excel_file(filepath)
    else:
        raise ValueError(f"Format file tidak didukung: {ext}")

    # Ekstrak hanya kolom yang relevan
    dataset = []
    for row in raw:
        # Cari kolom teks
        text_val = next((v for k, v in row.items()
                         if k.lower() == text_col.lower()), '')
        
        # Cari semua kolom label
        labels_dict = {}
        for lc in label_cols:
            val = next((v for k, v in row.items() if k.lower() == lc.lower()), '')
            labels_dict[lc] = val
        
        # Buat string representasi untuk display
        label_str = " | ".join(f"{k}: {v}" for k, v in labels_dict.items())
        
        dataset.append({'text': text_val, 'labels': labels_dict, 'label': label_str})

    return dataset


# BAGIAN 2: TEXT CLEANSING

# Pola Regex yang dikompilasi sekali untuk efisiensi (re.compile → O(1) lookup)
_RE_URL     = re.compile(r'https?://\S+|www\.\S+')
_RE_MENTION = re.compile(r'@\w+')
_RE_HASHTAG = re.compile(r'#\w+')
_RE_HTML    = re.compile(r'<[^>]+>')
_RE_NUMBER  = re.compile(r'\d+')
_RE_SYMBOL  = re.compile(r'[^a-zA-Z\s]')       # Hapus semua non-huruf non-spasi
_RE_SPACES  = re.compile(r'\s+')               # Normalisasi multi-spasi → satu spasi


def cleansing(text: str, steps_log: bool = False) -> str | tuple:
    if not isinstance(text, str):
        text = str(text)

    steps = []

    # Langkah 1: Lowercase — standarisasi huruf kapital
    t = text.lower()
    if steps_log: steps.append(('1. Lowercase', t))

    # Langkah 2: Hapus URL (http://, https://, www.)
    t = _RE_URL.sub('', t)
    if steps_log: steps.append(('2. Hapus URL', t))

    # Langkah 3: Hapus mention Twitter/sosmed (@username)
    t = _RE_MENTION.sub('', t)
    if steps_log: steps.append(('3. Hapus Mention', t))

    # Langkah 4: Hapus hashtag (#topik)
    t = _RE_HASHTAG.sub('', t)
    if steps_log: steps.append(('4. Hapus Hashtag', t))

    # Langkah 5: Hapus tag HTML (<br>, <p>, dll.)
    t = _RE_HTML.sub('', t)
    if steps_log: steps.append(('5. Hapus HTML', t))

    # Langkah 6: Hapus angka (digit [0-9])
    t = _RE_NUMBER.sub('', t)
    if steps_log: steps.append(('6. Hapus Angka', t))

    # Langkah 7: Hapus simbol & tanda baca → ganti dengan spasi
    # [^a-zA-Z\s] = karakter yang BUKAN huruf dan BUKAN whitespace
    t = _RE_SYMBOL.sub(' ', t)
    if steps_log: steps.append(('7. Hapus Simbol', t))

    # Langkah 8: Normalisasi spasi ganda → satu spasi, lalu strip tepi
    # re.sub(\s+) → menggabungkan semua whitespace berturut-turut
    t = _RE_SPACES.sub(' ', t).strip()
    if steps_log: steps.append(('8. Normalize Spasi', t))

    return (t, steps) if steps_log else t


# BAGIAN 3: SISTEM NORMALISASI LEKSIKON DINAMIS
def read_lexicon_file(
    filepath: str,
    slang_col: str = 'slang',
    formal_col: str = 'formal'
) -> dict:

    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.csv':
        rows = read_csv_file(filepath)
    elif ext in ('.xlsx', '.xls'):
        rows = read_excel_file(filepath)
    elif ext == '.json':
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, dict):
                return {str(k).lower().strip(): str(v).lower().strip() for k, v in data.items()}
            else:
                raise ValueError("Format JSON harus berupa objek/dictionary")
    else:
        raise ValueError(f"Format tidak didukung: {ext}")

    lexicon = {}
    for row in rows:
        # Case-insensitive column matching
        slang_val = next((v for k, v in row.items()
                          if k.lower() == slang_col.lower()), None)
        formal_val = next((v for k, v in row.items()
                           if k.lower() == formal_col.lower()), None)

        if slang_val and formal_val:
            # Normalisasi key: lowercase & strip untuk konsistensi lookup
            lexicon[slang_val.lower().strip()] = formal_val.lower().strip()

    return lexicon


def merge_lexicons(
    lexicon_internal: dict,
    lexicon_external: dict,
    conflict_strategy: str = 'internal_priority'
) -> dict:
    conflicts = {}   # {slang: {'internal': val, 'external': val}}
    duplicates = []  # kata yang sama DAN nilai yang sama (benar-benar duplikat)

    # Mulai dari salinan kamus internal (baseline)
    merged = dict(lexicon_internal)

    for slang, formal in lexicon_external.items():
        if slang in merged:
            if merged[slang] == formal:
                # Duplikat murni: kata & makna sama → abaikan
                duplicates.append(slang)
            else:
                # Konflik: kata sama tapi makna berbeda → catat & resolusi
                conflicts[slang] = {
                    'internal': merged[slang],
                    'external': formal,
                    'resolved': merged[slang]  # default: internal menang
                }
                if conflict_strategy == 'external_priority':
                    merged[slang] = formal
                    conflicts[slang]['resolved'] = formal
                # 'internal_priority' & 'keep_both' → pertahankan internal (sudah ada)
        else:
            # Kata baru dari eksternal → tambahkan
            merged[slang] = formal

    stats = {
        'total_internal'  : len(lexicon_internal),
        'total_external'  : len(lexicon_external),
        'total_merged'    : len(merged),
        'true_duplicates' : len(duplicates),
        'conflicts'       : conflicts,
        'conflict_count'  : len(conflicts),
        'new_from_external': len(merged) - len(lexicon_internal),
        'strategy_used'   : conflict_strategy
    }

    return merged, stats


def tokenize(text: str) -> list[str]:
    # str.split() tanpa argumen: split pada satu atau lebih whitespace
    return text.split()


def normalize_text(
    text: str,
    lexicon: dict,
    return_log: bool = False
) -> str | tuple:

    tokens = tokenize(text)  # Tokenisasi: O(n)
    normalized_tokens = []
    replacement_log = []

    for token in tokens:
        # Lookup O(1): cek apakah token ada di kamus
        if token in lexicon:
            replacement = lexicon[token]
            normalized_tokens.append(replacement)
            if return_log:
                replacement_log.append({
                    'original' : token,
                    'replaced' : replacement
                })
        else:
            # Token tidak ditemukan di kamus → pertahankan
            normalized_tokens.append(token)

    # Gabungkan token kembali menjadi kalimat
    result = ' '.join(normalized_tokens)

    return (result, replacement_log) if return_log else result

# BAGIAN 4: FULL PIPELINE
def full_preprocessing_pipeline(
    text: str,
    lexicon: dict,
    verbose: bool = False,
    use_stemming: bool = False,
    use_stopword: bool = False
) -> dict:
    # Tahap 1: Cleansing
    cleansed, cleansing_steps = cleansing(text, steps_log=True)

    # Tahap 2: Normalisasi
    normalized, replacement_log = normalize_text(
        cleansed, lexicon, return_log=True
    )

    # Tahap 3: Stopword Removal (Opsional)
    if use_stopword:
        sw_remover = get_stopword_remover()
        if sw_remover:
            normalized = sw_remover.remove(normalized)
            if verbose:
                cleansing_steps.append(('9. Stopword Removal', normalized))
        else:
            if verbose:
                cleansing_steps.append(('9. Stopword Removal', '[Lewati] Sastrawi tidak terinstall'))

    # Tahap 4: Stemming (Opsional)
    if use_stemming:
        stemmer = get_stemmer()
        if stemmer:
            normalized = stemmer.stem(normalized)
            if verbose:
                cleansing_steps.append(('10. Stemming', normalized))
        else:
            if verbose:
                cleansing_steps.append(('10. Stemming', '[Lewati] Sastrawi tidak terinstall'))

    result = {
        'raw'            : text,
        'cleansed'       : cleansed,
        'normalized'     : normalized,
        'cleansing_steps': cleansing_steps if verbose else [],
        'replacements'   : replacement_log,
        'stats': {
            'raw_length'       : len(text),
            'cleansed_length'  : len(cleansed),
            'normalized_length': len(normalized),
            'tokens_raw'       : len(tokenize(cleansed)),
            'tokens_normalized': len(tokenize(normalized)),
            'replacements_made': len(replacement_log)
        }
    }

    return result

# BAGIAN 5: KAMUS BAWAAN (FALLBACK JIKA TIDAK ADA FILE UPLOAD)
BUILTIN_INTERNAL_LEXICON = {
    
}

BUILTIN_EXTERNAL_LEXICON = {
    
}
